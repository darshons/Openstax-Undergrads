#!/usr/bin/env python3
"""Convert first_last_frame's output/first_last_frame/flf_log.json into a
labeling-ready Excel workbook. This pipeline has no auto-eval step (unlike
video_generator's scene pipeline) — each entry is just one clip's generation
attempt, so this is a single flat sheet keyed on generation success/failure
rather than eval pass/fail.

Usage:
    python scripts/flf_log_to_excel.py [path/to/flf_log.json] [path/to/output.xlsx]

Defaults to output/first_last_frame/flf_log.json -> output/first_last_frame/flf_log.xlsx.
"""
import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

NAVY = "1F3864"
LIGHT_BLUE = "DCE6F1"
WHITE = "FFFFFF"
RED = "FFC7CE"

HEADER_FONT = Font(color=WHITE, bold=True, size=11)
TITLE_FONT = Font(color=WHITE, bold=True, size=16)
HEADER_FILL = PatternFill("solid", fgColor=NAVY)
BAND_FILL = PatternFill("solid", fgColor=LIGHT_BLUE)
FAIL_FILL = PatternFill("solid", fgColor=RED)
THIN = Side(style="thin", color="B7C6DE")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def load_entries(path: Path) -> list:
    if not path.exists():
        return []
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def main():
    src = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("output/first_last_frame/flf_log.json")
    dst = Path(sys.argv[2]) if len(sys.argv) > 2 else src.with_suffix(".xlsx")

    entries = load_entries(src)

    wb = Workbook()
    ws = wb.active
    ws.title = "First-Last-Frame Log"

    ws.merge_cells("A1:M1")
    ws["A1"] = "First/Last-Frame Clip Generation Log"
    ws["A1"].font = TITLE_FONT
    ws["A1"].fill = HEADER_FILL
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    summary_row = 3
    table_header_row = 5
    first_data_row = table_header_row + 1
    last_data_row = first_data_row + max(len(entries), 1) - 1

    headers = [
        "Clip #", "Timestamp", "Scene", "Clip", "Model",
        "Generated", "Human Review", "Retries", "Duration (s)",
        "Generation Time (s)", "Cost (USD)", "Output File", "Error",
    ]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=table_header_row, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
    ws.row_dimensions[table_header_row].height = 30

    gen_col, review_col = "F", "G"
    time_col, cost_col = "J", "K"

    for idx, e in enumerate(entries):
        r = first_data_row + idx
        values = [
            idx + 1,
            e.get("timestamp"),
            e.get("scene_id"),
            e.get("clip_id"),
            e.get("model"),
            "Yes" if e.get("success") else "No",
            "Not Reviewed",
            e.get("retry_count", 0),
            e.get("duration_seconds"),
            e.get("time_seconds"),
            e.get("cost_usd"),
            e.get("output_file"),
            e.get("error"),
        ]
        for col, v in enumerate(values, start=1):
            cell = ws.cell(row=r, column=col, value=v)
            cell.border = BORDER
            if not e.get("success"):
                cell.fill = FAIL_FILL
            elif idx % 2 == 1:
                cell.fill = BAND_FILL

    dv_yesno = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    dv_review = DataValidation(type="list", formula1='"Not Reviewed,Approved,Rejected,Needs Rework"', allow_blank=True)
    for dv in (dv_yesno, dv_review):
        ws.add_data_validation(dv)
    dv_yesno.add(f"{gen_col}{first_data_row}:{gen_col}{last_data_row}")
    dv_review.add(f"{review_col}{first_data_row}:{review_col}{last_data_row}")

    ws.cell(row=summary_row, column=1, value="Clips").font = Font(bold=True)
    ws.cell(row=summary_row, column=2, value=f"=COUNTA(A{first_data_row}:A{last_data_row})")
    ws.cell(row=summary_row, column=4, value="Gen. Success").font = Font(bold=True)
    ws.cell(row=summary_row, column=5,
            value=f'=COUNTIF({gen_col}{first_data_row}:{gen_col}{last_data_row},"Yes")/COUNTA(A{first_data_row}:A{last_data_row})')
    ws.cell(row=summary_row, column=5).number_format = "0.0%"
    ws.cell(row=summary_row, column=7, value="Avg Time (s)").font = Font(bold=True)
    ws.cell(row=summary_row, column=8, value=f"=AVERAGE({time_col}{first_data_row}:{time_col}{last_data_row})")
    ws.cell(row=summary_row, column=10, value="Total Cost ($)").font = Font(bold=True)
    ws.cell(row=summary_row, column=11, value=f"=SUM({cost_col}{first_data_row}:{cost_col}{last_data_row})")
    ws.cell(row=summary_row, column=11).number_format = '"$"#,##0.00'

    ws.freeze_panes = f"A{first_data_row}"
    ws.auto_filter.ref = f"A{table_header_row}:M{max(last_data_row, table_header_row)}"

    widths = [7, 20, 7, 6, 14, 10, 15, 8, 12, 17, 11, 55, 30]
    for col, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = w

    wb.save(dst)
    print(f"Wrote {dst} ({len(entries)} clips)")

    n = len(entries)
    failed = sum(1 for e in entries if not e.get("success"))
    if n:
        print(f"Generation failure rate: {failed}/{n} ({failed/n:.1%})")
    print("Note: this pipeline has no auto-eval step, so 'eval rejection rate' doesn't "
          "apply here — that stat lives in generation_log_to_excel.py's Clips sheet.")


if __name__ == "__main__":
    main()
