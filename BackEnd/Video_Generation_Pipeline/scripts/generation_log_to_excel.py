#!/usr/bin/env python3
"""Convert video_generator's output/generation_log.json into a labeling-ready
Excel workbook: one "Scenes" sheet (one row per scene-attempt) plus a
"Clips (Eval Detail)" sheet (one row per individual clip's auto-eval result,
since eval runs and rejects at the clip level, not the scene level) with the
rejection-rate / wasted-cost summary used to estimate eval overhead cost.

Usage:
    python scripts/generation_log_to_excel.py [path] [path/to/output.xlsx]

[path] may point directly at a generation_log.json, or at an experiment
folder (e.g. output/experiments/my_experiment_20260724_090818/) — each
run_experiment.py run gets its own such folder with a generation_log.json
inside it, and this resolves that automatically.

Defaults to output/generation_log.json -> output/generation_log.xlsx.
"""
import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

NAVY = "1F3864"
LIGHT_BLUE = "DCE6F1"
WHITE = "FFFFFF"
RED = "FFC7CE"

HEADER_FONT = Font(color=WHITE, bold=True, size=11)
TITLE_FONT = Font(color=WHITE, bold=True, size=16)
HEADER_FILL = PatternFill("solid", fgColor=NAVY)
BAND_FILL = PatternFill("solid", fgColor=LIGHT_BLUE)
FAIL_FILL = PatternFill("solid", fgColor=RED)
THIN = Side(style="thin", color="B7C6DE")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def load_entries(path: Path) -> list:
    if not path.exists():
        return []
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def resolve_source(path: Path) -> Path:
    """Accept either a generation_log.json file or an experiment folder
    (output/experiments/<name>_<timestamp>/) containing one."""
    if path.is_dir():
        log_path = path / "generation_log.json"
        if not log_path.exists():
            raise SystemExit(f"No generation_log.json found in experiment folder: {path}")
        return log_path
    return path


def flatten_clips(entries: list) -> list:
    rows = []
    for e in entries:
        for i, c in enumerate(e.get("clips") or [], start=1):
            rows.append({
                "scene_id": e.get("scene_id"),
                "scene_attempt": e.get("scene_attempt"),
                "clip_index": i,
                "eval_passed": c.get("eval_passed"),
                "estimated_cost_usd": c.get("estimated_cost_usd"),
                "error": c.get("error"),
            })
    return rows


def derive_meta(entries: list, src: Path, clip_rows: list) -> dict:
    """Pull the fields that are constant across a single experiment run
    straight out of the log, instead of leaving them for manual entry."""
    experiment_name = src.parent.name if src.parent.name != "output" else ""

    timestamps = sorted(e["timestamp"] for e in entries if e.get("timestamp"))
    experiment_date = timestamps[0][:10] if timestamps else ""

    models = sorted({e.get("model") for e in entries if e.get("model")})
    primary_model = ", ".join(models) if models else ""

    retry_counts = sorted({e.get("retry_count") for e in entries if e.get("retry_count") is not None})
    if len(retry_counts) == 1:
        retry_limit = str(retry_counts[0])
    elif retry_counts:
        retry_limit = f"varies ({retry_counts[0]}-{retry_counts[-1]})"
    else:
        retry_limit = ""

    opening_prompts = set()
    for e in entries:
        clips = e.get("clips") or []
        if clips and clips[0].get("prompt"):
            opening_prompts.add(clips[0]["prompt"])
    if len(opening_prompts) == 1:
        prompt = next(iter(opening_prompts))
    elif opening_prompts:
        prompt = f"(varies across {len(opening_prompts)} scenes/prompts — see clips[].prompt in the source log)"
    else:
        prompt = ""

    failed_clips = sum(1 for r in clip_rows if r["eval_passed"] is False)

    return {
        "experiment_name": experiment_name,
        "experiment_date": experiment_date,
        "primary_model": primary_model,
        "retry_limit": retry_limit,
        "prompt": prompt,
        "clips_rejected_by_eval": failed_clips,
    }


def build_scenes_sheet(wb, entries, meta=None):
    ws = wb.active
    ws.title = "Scenes"

    ws.merge_cells("A1:P1")
    ws["A1"] = "Video Generator Experiment Log"
    ws["A1"].font = TITLE_FONT
    ws["A1"].fill = HEADER_FILL
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    meta = meta or {}
    meta_fields = [
        ("Experiment Name", meta.get("experiment_name", "")),
        ("Test Purpose", ""),
        ("Owner", ""),
        ("Experiment Date", meta.get("experiment_date", "")),
        ("Primary Model", meta.get("primary_model", "")),
        ("Retry Limit", meta.get("retry_limit", "")),
        ("Prompt / Scenario File", meta.get("prompt", "")),
        ("Clips Rejected by Eval", meta.get("clips_rejected_by_eval", "")),
        ("Experiment Notes", ""),
    ]
    row = 3
    for label, value in meta_fields:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        cell = ws.cell(row=row, column=2, value=value)
        if label == "Prompt / Scenario File":
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[row].height = 60
        row += 1

    summary_row = row + 1
    table_header_row = summary_row + 2
    first_data_row = table_header_row + 1
    last_data_row = first_data_row + max(len(entries), 1) - 1

    headers = [
        "Video #", "Timestamp", "Scene", "Scene Attempt", "Model", "Clips",
        "Generated", "Auto Eval", "Human Review", "Retries",
        "Video Duration (s)", "Generation Time (s)", "Cost (USD)",
        "File Size (MB)", "Output File", "Error",
    ]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=table_header_row, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
    ws.row_dimensions[table_header_row].height = 30

    gen_col, eval_col, review_col = "G", "H", "I"
    time_col, cost_col = "L", "M"

    for idx, e in enumerate(entries):
        r = first_data_row + idx
        clips = e.get("clips") or []
        values = [
            idx + 1,
            e.get("timestamp"),
            e.get("scene_id"),
            e.get("scene_attempt"),
            e.get("model"),
            len(clips),
            "Yes" if e.get("success") else "No",
            {True: "Pass", False: "Fail", None: "Not Run"}.get(e.get("eval_passed")),
            "Not Reviewed",
            e.get("retry_count", 0),
            e.get("final_video_duration_seconds"),
            e.get("total_time_seconds"),
            e.get("total_cost_usd"),
            e.get("final_file_size_mb"),
            e.get("final_output_file"),
            e.get("error"),
        ]
        for col, v in enumerate(values, start=1):
            cell = ws.cell(row=r, column=col, value=v)
            cell.border = BORDER
            if not e.get("success"):
                cell.fill = FAIL_FILL
            elif idx % 2 == 1:
                cell.fill = BAND_FILL

    dv_yesno = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    dv_eval = DataValidation(type="list", formula1='"Pass,Fail,Not Run"', allow_blank=True)
    dv_review = DataValidation(type="list", formula1='"Not Reviewed,Approved,Rejected,Needs Rework"', allow_blank=True)
    for dv in (dv_yesno, dv_eval, dv_review):
        ws.add_data_validation(dv)
    dv_yesno.add(f"{gen_col}{first_data_row}:{gen_col}{last_data_row}")
    dv_eval.add(f"{eval_col}{first_data_row}:{eval_col}{last_data_row}")
    dv_review.add(f"{review_col}{first_data_row}:{review_col}{last_data_row}")

    ws.cell(row=summary_row, column=1, value="Videos").font = Font(bold=True)
    ws.cell(row=summary_row, column=2, value=f"=COUNTA(A{first_data_row}:A{last_data_row})")
    ws.cell(row=summary_row, column=4, value="Gen. Success").font = Font(bold=True)
    ws.cell(row=summary_row, column=5,
            value=f'=COUNTIF({gen_col}{first_data_row}:{gen_col}{last_data_row},"Yes")/COUNTA(A{first_data_row}:A{last_data_row})')
    ws.cell(row=summary_row, column=5).number_format = "0.0%"
    ws.cell(row=summary_row, column=7, value="Avg Time (s)").font = Font(bold=True)
    ws.cell(row=summary_row, column=8, value=f"=AVERAGE({time_col}{first_data_row}:{time_col}{last_data_row})")
    ws.cell(row=summary_row, column=10, value="Total Cost ($)").font = Font(bold=True)
    ws.cell(row=summary_row, column=11, value=f"=SUM({cost_col}{first_data_row}:{cost_col}{last_data_row})")
    ws.cell(row=summary_row, column=11).number_format = '"$"#,##0.00'

    ws.freeze_panes = f"A{first_data_row}"
    ws.auto_filter.ref = f"A{table_header_row}:P{max(last_data_row, table_header_row)}"

    widths = [8, 20, 7, 14, 14, 7, 11, 10, 15, 8, 16, 17, 11, 12, 55, 30]
    for col, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = w


def build_clips_sheet(wb, rows):
    ws = wb.create_sheet("Clips (Eval Detail)")

    ws.merge_cells("A1:F1")
    ws["A1"] = "Per-Clip Eval Results"
    ws["A1"].font = TITLE_FONT
    ws["A1"].fill = HEADER_FILL
    ws.row_dimensions[1].height = 24

    header_row = 3
    first_data_row = header_row + 1
    last_data_row = first_data_row + max(len(rows), 1) - 1

    headers = ["Scene", "Scene Attempt", "Clip #", "Eval Result", "Est. Cost (USD)", "Error"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=header_row, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.border = BORDER

    for idx, r in enumerate(rows):
        row_num = first_data_row + idx
        ep = r["eval_passed"]
        values = [
            r["scene_id"], r["scene_attempt"], r["clip_index"],
            {True: "Pass", False: "Fail", None: "N/A"}.get(ep),
            r["estimated_cost_usd"], r["error"],
        ]
        for col, v in enumerate(values, start=1):
            cell = ws.cell(row=row_num, column=col, value=v)
            cell.border = BORDER
            if ep is False:
                cell.fill = FAIL_FILL
            elif idx % 2 == 1:
                cell.fill = BAND_FILL

    evaluated = [r for r in rows if r["eval_passed"] is not None]
    failed = [r for r in evaluated if r["eval_passed"] is False]
    cost_failed = sum(r["estimated_cost_usd"] or 0 for r in failed)
    pct = (len(failed) / len(evaluated)) if evaluated else 0

    ws.cell(row=1, column=8, value="Clips Evaluated").font = Font(bold=True)
    ws.cell(row=1, column=9, value=len(evaluated))
    ws.cell(row=2, column=8, value="Rejected by Eval").font = Font(bold=True)
    ws.cell(row=2, column=9, value=len(failed))
    ws.cell(row=1, column=10, value="Rejection Rate").font = Font(bold=True)
    ws.cell(row=1, column=11, value=pct).number_format = "0.0%"
    ws.cell(row=2, column=10, value="Cost Burned on Rejects ($)").font = Font(bold=True)
    ws.cell(row=2, column=11, value=round(cost_failed, 4)).number_format = '"$"#,##0.00'

    ws.freeze_panes = f"A{first_data_row}"
    ws.auto_filter.ref = f"A{header_row}:F{max(last_data_row, header_row)}"
    widths = [7, 14, 8, 12, 15, 40]
    for col, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = w

    return pct, len(failed), len(evaluated), cost_failed


def main():
    arg = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("output/generation_log.json")
    src = resolve_source(arg)
    dst = Path(sys.argv[2]) if len(sys.argv) > 2 else src.with_suffix(".xlsx")

    entries = load_entries(src)
    scene_entries = [e for e in entries if e.get("type") != "checkpoint"]
    clip_rows = flatten_clips(scene_entries)
    meta = derive_meta(scene_entries, src, clip_rows)

    wb = Workbook()
    build_scenes_sheet(wb, scene_entries, meta)
    pct, failed, evaluated, cost = build_clips_sheet(wb, clip_rows)
    wb.save(dst)

    print(f"Wrote {dst} ({len(scene_entries)} scene-attempts)")
    if evaluated:
        print(f"Eval rejection rate: {failed}/{evaluated} clips ({pct:.1%}), "
              f"${cost:.2f} spent generating clips the eval system rejected")
    else:
        print("No clips have eval results yet (source log is empty or verify_clips=False).")


if __name__ == "__main__":
    main()
